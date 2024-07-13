import json
import openpyxl
from openpyxl.cell import Cell
from apps.staff.models import Staff
from apps.faculties.models import Faculty
from apps.specialties.models import Specialty


def read_initial_data() -> dict[str, dict]:
    """
    read initial data from its json file
    """
    with open('scripts/inputs/initial_data.json', 'r', encoding='utf-8') as file:
      return json.loads(file.read())


def db_reset() -> None:
    """
    delete all records in the database
    """
    Staff.objects.all().delete()
    Specialty.objects.all().delete()
    Faculty.objects.all().delete()
    
    print('done reset! 😁')


def db_import_faculties_specialties() -> None:
    """
    import the default faculties and specialties
    """
    initial_data = read_initial_data()
    
    specialties: list[Specialty] = []
    
    for faculty in initial_data.values():
        name = faculty['name']
        faculty_instance = Faculty.objects.create(name=name)
        
        specialties_initial: list[dict] = faculty['specialties']
        for specialty in specialties_initial:
            s = Specialty(faculty=faculty_instance, **specialty)
            specialties.append(s)
            
    Specialty.objects.bulk_create(specialties)
    
    print('done importing faculties and specialties! 😁')


def read_xlsx_data(path: str) -> list[dict]:
    """
    import staff from staff.xlsx file
    """
    data: list[dict] = []
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    lr = ws.max_row
    
    headers_rg: list[Cell] = ws[f'A1:F1'][0]
    headers = [cell.value for cell in headers_rg]
    
    for idx in range(2, lr):
        row_rg: list[Cell] = ws[f'A{idx}:F{idx}'][0]
        row = {k: v.value for k, v in zip(headers, row_rg)}
        data.append(row)
      
    return data


def db_import_staff() -> None:
    """
    import staff from staff.xlsx file
    """
    xlsx_path:str  = 'scripts/inputs/staff.xlsx'
    staff_data = read_xlsx_data(xlsx_path)
    
    initial_data = read_initial_data()
    staff_list: list[Staff] = []
    
    for staff in staff_data:

        faculty_name: str = initial_data[staff['faculty']]['name']
        faculty_instance = Faculty.objects.get(name=faculty_name)
        
        specialty_instance = (
          Specialty
          .objects
          .get(faculty=faculty_instance, name=staff['specialty'])
        )
        
        del staff['faculty']
        staff['specialty'] = specialty_instance
        
        staff_list.append(Staff(**staff))
        
    Staff.objects.bulk_create(staff_list)
        
    print('done importing staff! 😁')